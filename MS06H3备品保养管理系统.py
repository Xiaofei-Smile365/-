# -*- coding: UTF-8 -*-

"""

@author:smile

@file:MS06H3备品保养管理系统.py

@time:2020/04/10

"""

import sys  # 载入必需的模块
import os
import shutil

if hasattr(sys, 'frozen'):
    os.environ['PATH'] = sys._MEIPASS + ";" + os.environ['PATH']

from PyQt5.QtGui import *
from PyQt5.QtWidgets import *  # 在Qt5中使用的基本的GUI窗口控件都在PyQt5.QtWidgets模块中
import pandas as pd
from playsound import playsound
from PyQt5 import QtCore
from openpyxl import load_workbook
import datetime


def alarm_sound():

    playsound('./Alarm_Sound.mp3')


def need_care_tool(self):

    def care_write(self, row, toolname, lastcaredata, careperson, carecycle, nextcaredata, surplusdays, aheaddays, isalarm):
        newItem_toolname = QTableWidgetItem(str(toolname))  # 添加数据
        self.table_view.setItem(row, 0, newItem_toolname)
        newItem_lastcaredata = QTableWidgetItem(str(lastcaredata)[0:10])
        self.table_view.setItem(row, 1, newItem_lastcaredata)
        newItem_careperson = QTableWidgetItem(str(careperson))
        self.table_view.setItem(row, 2, newItem_careperson)
        newItem_carecycle = QTableWidgetItem(str(carecycle)[:-2])
        self.table_view.setItem(row, 3, newItem_carecycle)
        newItem_nextcaredata = QTableWidgetItem(str(nextcaredata)[0:10])
        self.table_view.setItem(row, 4, newItem_nextcaredata)
        newItem_surplusdays = QTableWidgetItem(str(surplusdays)[:-2])
        self.table_view.setItem(row, 5, newItem_surplusdays)
        newItem_aheaddays = QTableWidgetItem(str(aheaddays)[:-2])
        self.table_view.setItem(row, 6, newItem_aheaddays)
        newItem_isalarm = QTableWidgetItem(str(isalarm)[:-2])
        self.table_view.setItem(row, 7, newItem_isalarm)

    def care_write_all_data(self, row, toolname, lastcaredata, careperson, carecycle, nextcaredata, surplusdays, aheaddays, isalarm):
        newItem_toolname = QTableWidgetItem(str(toolname))  # 添加数据
        self.table_view_all_data.setItem(row, 0, newItem_toolname)
        newItem_lastcaredata = QTableWidgetItem(str(lastcaredata)[0:10])
        self.table_view_all_data.setItem(row, 1, newItem_lastcaredata)
        newItem_careperson = QTableWidgetItem(str(careperson))
        self.table_view_all_data.setItem(row, 2, newItem_careperson)
        newItem_carecycle = QTableWidgetItem(str(carecycle)[:-2])
        self.table_view_all_data.setItem(row, 3, newItem_carecycle)
        newItem_nextcaredata = QTableWidgetItem(str(nextcaredata)[0:10])
        self.table_view_all_data.setItem(row, 4, newItem_nextcaredata)
        newItem_surplusdays = QTableWidgetItem(str(surplusdays)[:-2])
        self.table_view_all_data.setItem(row, 5, newItem_surplusdays)
        newItem_aheaddays = QTableWidgetItem(str(aheaddays)[:-2])
        self.table_view_all_data.setItem(row, 6, newItem_aheaddays)
        newItem_isalarm = QTableWidgetItem(str(isalarm)[:-2])
        self.table_view_all_data.setItem(row, 7, newItem_isalarm)

    def if_create_new_source_file(new_file):
        if not os.path.exists(new_file):  # 判断新文件是否存在，否则就创建该文件
            file = open(new_file, 'w')
            new_file_columns = ['备品名称（手动写入）', '上次保养日期（手动写入）', '保养人员（手动写入）', '保养周期（手动写入）', '下次保养日期（自动计算）',
                                '距离下次保养日期剩余天数（自动计算）', '提前预警天数（手动写入）', '是否报警（<0为报警，>0为安全；自动计算，剩余天数-预警天数）']
            data = pd.DataFrame(columns=new_file_columns)
            data.to_excel(new_file, index=False)
            file.close()

    sourcefile = './备品保养数据.xlsx'  # 获取数据源
    if_create_new_source_file(sourcefile)
    self.table_view.clearContents()
    sourcefile_frame = pd.DataFrame(pd.read_excel(sourcefile))
    sourcefile_need_care_ahead_days = sourcefile_frame[sourcefile_frame['是否报警（<0为报警，>0为安全；自动计算，剩余天数-预警天数）'] <= 0]  # 获取预警的数据
    sourcefile_need_care_ahead_days = sourcefile_need_care_ahead_days[sourcefile_need_care_ahead_days['距离下次保养日期剩余天数（自动计算）'] > 0]
    sourcefile_need_care_today = sourcefile_frame[sourcefile_frame['距离下次保养日期剩余天数（自动计算）'] == 0]  # 获取当天必须保养的数据
    sourcefile_need_care_ng = sourcefile_frame[pd.to_numeric(sourcefile_frame['距离下次保养日期剩余天数（自动计算）']) < 0]  # 获取过保养日期的数据

    need_care_all_list = []
    for i in range(0, len(sourcefile_frame)):
        null = 'nan'
        if str(sourcefile_frame.iat[i, 0]) != str(null):
            for j in range(0, 8):
                need_care_all_list.append(sourcefile_frame.iat[i, j])
            care_write_all_data(self, i, need_care_all_list[(i*8)+0], need_care_all_list[(i*8)+1], need_care_all_list[(i*8)+2], need_care_all_list[(i*8)+3], need_care_all_list[(i*8)+4], need_care_all_list[(i*8)+5], need_care_all_list[(i*8)+6], need_care_all_list[(i*8)+7])

    need_care_ng_list = []
    for i in range(0, len(sourcefile_need_care_ng)):
        for j in range(0, 8):
            need_care_ng_list.append(sourcefile_need_care_ng.iat[i, j])
        care_write(self, i, need_care_ng_list[(i*8)+0], need_care_ng_list[(i*8)+1], need_care_ng_list[(i*8)+2], need_care_ng_list[(i*8)+3], need_care_ng_list[(i*8)+4], need_care_ng_list[(i*8)+5], need_care_ng_list[(i*8)+6], need_care_ng_list[(i*8)+7])
    if len(sourcefile_need_care_ng) != 0:
        self.label_alarm.setPixmap(QPixmap('./alarm_red.png'))
        playsound('./Alarm_Sound.mp3')
        self.time_1m = QtCore.QTimer()
        self.time_1m.timeout.connect(alarm_sound)
        self.time_1m.start(60000)
    elif len(sourcefile_need_care_ng) == 0:
        self.label_alarm.setPixmap(QPixmap('./alarm_green.png'))
        try:
            self.time_1m.stop()
        except:
            pass

    need_care_today_list = []
    for i in range(0, len(sourcefile_need_care_today)):
        for j in range(0, 8):
            need_care_today_list.append(sourcefile_need_care_today.iat[i, j])
        care_write(self, i+len(sourcefile_need_care_ng), need_care_today_list[(i*8)+0], need_care_today_list[(i*8)+1], need_care_today_list[(i*8)+2], need_care_today_list[(i*8)+3], need_care_today_list[(i*8)+4], need_care_today_list[(i*8)+5], need_care_today_list[(i*8)+6], need_care_today_list[(i*8)+7])
    if len(sourcefile_need_care_today) != 0:
        self.label_alarm.setPixmap(QPixmap('./alarm_red.png'))
        playsound('./Alarm_Sound.mp3')
        self.time_30m = QtCore.QTimer()
        self.time_30m.timeout.connect(alarm_sound)
        self.time_30m.start(1800000)
    elif len(sourcefile_need_care_today) == 0 and len(sourcefile_need_care_ng) == 0:
        self.label_alarm.setPixmap(QPixmap('./alarm_green.png'))
        try:
            self.time_30m.stop()
        except:
            pass

    need_care_ahead_days_list = []
    for i in range(0, len(sourcefile_need_care_ahead_days)):
        for j in range(0, 8):
            need_care_ahead_days_list.append(sourcefile_need_care_ahead_days.iat[i, j])
        care_write(self, i+len(sourcefile_need_care_ng)+len(sourcefile_need_care_today), need_care_ahead_days_list[(i*8)+0], need_care_ahead_days_list[(i*8)+1], need_care_ahead_days_list[(i*8)+2], need_care_ahead_days_list[(i*8)+3], need_care_ahead_days_list[(i*8)+4], need_care_ahead_days_list[(i*8)+5], need_care_ahead_days_list[(i*8)+6], need_care_ahead_days_list[(i*8)+7])
    if len(sourcefile_need_care_ahead_days) != 0:
        self.label_alarm.setPixmap(QPixmap('./alarm_red.png'))
        playsound('./Alarm_Sound.mp3')
        self.time_120m = QtCore.QTimer()
        self.time_120m.timeout.connect(alarm_sound)
        self.time_120m.start(7200000)
    elif len(sourcefile_need_care_ahead_days) == 0 and len(sourcefile_need_care_today) == 0 and len(sourcefile_need_care_ng) == 0:
        self.label_alarm.setPixmap(QPixmap('./alarm_green.png'))
        try:
            self.time_120m.stop()
        except:
            pass

    path = os.path.dirname(os.getcwd())
    copy_path = str(str(path).replace("\\", "/") + '/备品保养数据' + ".xlsx")
    shutil.copyfile(sourcefile, copy_path)


class MainWindow(QWidget):
    def __init__(self, parent=None):  # 基础窗口控件QWidget类是所有用户界面对象的基类， 所有的窗口和控件都直接或间接继承自QWidget类。
        super(MainWindow, self).__init__(parent)  # 使用super函数初始化窗口

        self.setWindowTitle('MS06H3 备品保养管理系统')  # 设定窗口控件的标题
        self.setWindowIcon(QIcon('./tool.ico'))  # 设定窗口的图标

        self.move(100, 100)

        self.label_title = QLabel(self)
        self.label_title.setText('<b>MS06H3 备品保养管理系统<b>')
        self.label_title.setFont(QFont('SanSerif', 16))
        self.label_title.setFixedSize(258, 20)

        self.table_view = QTableWidget()
        self.table_view.setRowCount(99)
        self.table_view.setColumnCount(8)
        self.table_view.setFixedSize(1475, 150)
        self.table_view.setHorizontalHeaderLabels(['备品名称（手动写入）', '上次保养日期（手动写入）', '保养人员（手动写入）', '保养周期（手动写入）', '下次保养日期（自动计算）', '距离下次保养日期剩余天数（自动计算）', '提前预警天数（手动写入）', '是否报警（<0为报警，>0为安全；自动计算，剩余天数-预警天数）'])

        self.table_view.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_view.setSelectionBehavior(QAbstractItemView.SelectRows)
        QTableWidget.resizeColumnsToContents(self.table_view)
        QTableWidget.resizeRowsToContents(self.table_view)

        self.table_view_all_data = QTableWidget()
        self.table_view_all_data.setRowCount(999)
        self.table_view_all_data.setColumnCount(8)
        self.table_view_all_data.setFixedSize(1481, 252)
        self.table_view_all_data.setHorizontalHeaderLabels(['备品名称（手动写入）', '上次保养日期（手动写入）', '保养人员（手动写入）', '保养周期（手动写入）', '下次保养日期（自动计算）', '距离下次保养日期剩余天数（自动计算）', '提前预警天数（手动写入）', '是否报警（<0为报警，>0为安全；自动计算，剩余天数-预警天数）'])

        self.table_view_all_data.setEditTriggers(QAbstractItemView.NoEditTriggers)
        self.table_view_all_data.setSelectionBehavior(QAbstractItemView.SelectRows)
        QTableWidget.resizeColumnsToContents(self.table_view_all_data)
        QTableWidget.resizeRowsToContents(self.table_view_all_data)

        self.label_tool_name = QLabel(self)
        self.label_tool_name.setText('备品名称')
        self.label_tool_name.setFont(QFont('SanSerif', 10))
        self.label_tool_name.setFixedSize(55, 15)
        h_box_label_tool_name = QHBoxLayout()
        h_box_label_tool_name.addWidget(self.label_tool_name)

        self.label_tool_name_sample = QLabel(self)
        self.label_tool_name_sample.setText('示例：H139CLN02.2-OTP-01')
        self.label_tool_name_sample.setFont(QFont('SanSerif', 10))
        self.label_tool_name_sample.setFixedSize(163, 15)
        h_box_label_tool_name_sample = QHBoxLayout()
        h_box_label_tool_name_sample.addWidget(self.label_tool_name_sample)

        self.edit_tool_name = QLineEdit(self)
        self.edit_tool_name.setFixedSize(220, 30)
        h_box_edit_tool_name = QHBoxLayout()
        h_box_edit_tool_name.addWidget(self.edit_tool_name)

        v_box_tool_name = QVBoxLayout()
        v_box_tool_name.addLayout(h_box_label_tool_name)
        v_box_tool_name.addLayout(h_box_label_tool_name_sample)
        v_box_tool_name.addLayout(h_box_edit_tool_name)

        self.label_care_date = QLabel(self)
        self.label_care_date.setText('保养日期')
        self.label_care_date.setFont(QFont('SanSerif', 10))
        self.label_care_date.setFixedSize(55, 15)
        h_box_label_care_date = QHBoxLayout()
        h_box_label_care_date.addWidget(self.label_care_date)

        self.label_care_date_sample = QLabel(self)
        self.label_care_date_sample.setText('示例：2020/05/01')
        self.label_care_date_sample.setFont(QFont('SanSerif', 10))
        self.label_care_date_sample.setFixedSize(108, 15)
        h_box_label_care_date_sample = QHBoxLayout()
        h_box_label_care_date_sample.addWidget(self.label_care_date_sample)

        self.edit_care_date = QLineEdit(self)
        self.edit_care_date.setFixedSize(220, 30)
        h_box_edit_care_date = QHBoxLayout()
        h_box_edit_care_date.addWidget(self.edit_care_date)

        v_box_care_date = QVBoxLayout()
        v_box_care_date.addLayout(h_box_label_care_date)
        v_box_care_date.addLayout(h_box_label_care_date_sample)
        v_box_care_date.addLayout(h_box_edit_care_date)

        self.label_care_person = QLabel(self)
        self.label_care_person.setText('保养人员')
        self.label_care_person.setFont(QFont('SanSerif', 10))
        self.label_care_person.setFixedSize(55, 15)
        h_box_label_care_person = QHBoxLayout()
        h_box_label_care_person.addWidget(self.label_care_person)

        self.label_care_person_sample = QLabel(self)
        self.label_care_person_sample.setText('示例：S1710009 苏晓飞')
        self.label_care_person_sample.setFont(QFont('SanSerif', 10))
        self.label_care_person_sample.setFixedSize(140, 15)
        h_box_label_care_person_sample= QHBoxLayout()
        h_box_label_care_person_sample.addWidget(self.label_care_person_sample)

        self.edit_care_person = QLineEdit(self)
        self.edit_care_person.setFixedSize(220, 30)
        h_box_edit_care_person = QHBoxLayout()
        h_box_edit_care_person.addWidget(self.edit_care_person)

        v_box_care_person = QVBoxLayout()
        v_box_care_person.addLayout(h_box_label_care_person)
        v_box_care_person.addLayout(h_box_label_care_person_sample)
        v_box_care_person.addLayout(h_box_edit_care_person)

        self.label_care_cycle = QLabel(self)
        self.label_care_cycle.setText('保养周期')
        self.label_care_cycle.setFont(QFont('SanSerif', 10))
        self.label_care_cycle.setFixedSize(55, 15)
        h_box_label_care_cycle = QHBoxLayout()
        h_box_label_care_cycle.addWidget(self.label_care_cycle)

        self.label_care_cycle_sample = QLabel(self)
        self.label_care_cycle_sample.setText('示例：30')
        self.label_care_cycle_sample.setFont(QFont('SanSerif', 10))
        self.label_care_cycle_sample.setFixedSize(52, 15)
        h_box_label_care_cycle_sample= QHBoxLayout()
        h_box_label_care_cycle_sample.addWidget(self.label_care_cycle_sample)

        self.edit_care_cycle = QLineEdit(self)
        self.edit_care_cycle.setFixedSize(220, 30)
        h_box_edit_care_cycle = QHBoxLayout()
        h_box_edit_care_cycle.addWidget(self.edit_care_cycle)

        v_box_care_cycle = QVBoxLayout()
        v_box_care_cycle.addLayout(h_box_label_care_cycle)
        v_box_care_cycle.addLayout(h_box_label_care_cycle_sample)
        v_box_care_cycle.addLayout(h_box_edit_care_cycle)

        self.label_warning_days = QLabel(self)
        self.label_warning_days.setText('预警天数')
        self.label_warning_days.setFont(QFont('SanSerif', 10))
        self.label_warning_days.setFixedSize(55, 15)
        h_box_label_warning_days = QHBoxLayout()
        h_box_label_warning_days.addWidget(self.label_warning_days)

        self.label_warning_days_sample = QLabel(self)
        self.label_warning_days_sample.setText('示例：3')
        self.label_warning_days_sample.setFont(QFont('SanSerif', 10))
        self.label_warning_days_sample.setFixedSize(45, 15)
        h_box_label_warning_days_sample = QHBoxLayout()
        h_box_label_warning_days_sample.addWidget(self.label_warning_days_sample)

        self.edit_warning_days = QLineEdit(self)
        self.edit_warning_days.setFixedSize(220, 30)
        h_box_edit_warning_days = QHBoxLayout()
        h_box_edit_warning_days.addWidget(self.edit_warning_days)

        v_box_warning_days = QVBoxLayout()
        v_box_warning_days.addLayout(h_box_label_warning_days)
        v_box_warning_days.addLayout(h_box_label_warning_days_sample)
        v_box_warning_days.addLayout(h_box_edit_warning_days)

        self.label_alarm = QLabel(self)
        pix = QPixmap('./alarm_green.png')
        self.label_alarm.setPixmap(pix)
        self.label_alarm.setScaledContents(True)
        self.label_alarm.setFixedSize(200, 200)
        h_box_label_alarm = QHBoxLayout()
        h_box_label_alarm.addWidget(self.label_alarm)

        self.button_care_finish = QPushButton()
        self.button_care_finish.setText('写入此条治具保养信息')
        self.button_care_finish.setFont(QFont('SanSerif', 20))
        self.button_care_finish.setFixedSize(1470, 50)
        self.button_care_finish.setIcon(QIcon('./write.ico'))
        self.button_care_finish.pressed.connect(self.button_color_green)
        self.button_care_finish.released.connect(self.button_color_white)
        self.button_care_finish.clicked.connect(self.care_tool_write)  # 按钮被点击&释放后触发该信号
        h_box_button_care_finish = QHBoxLayout()
        h_box_button_care_finish.addWidget(self.button_care_finish)

        self.label_ps = QLabel(self)
        ps_txt = '1. 系统原理：此管理系统内部记录相关治具的保养信息，定时(2h)侦测距离下次保养日期的剩余天数，当达到预警天数时，报警提醒人员进行治具保养，同时在下方表格处列出需保养的治具名单。\n' \
                 '2. 操作指南：按照示例格式写入备品名称（同实际治具名称）、保养日期（格式务必为****/**/**）、保养人员（工号 姓名）、保养周期（格式为纯数字）、预警天数（格式为纯数字）信息后，点击下方按钮【写入此条治具保养信息】录入到数据库；\n' \
                 '3. 报警规则：当治具达到预警天数时，每2h报警一次；距离保养日期仅剩1天时，每0.5h报警一次；存在治具过保养日期后，每分钟报警一次直到该治具完成保养任务。\n' \
                 '------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------\n' \
                 '如有疑问，敬请联系！                                                                                ' \
                 '设计者：苏晓飞                                                                                ' \
                 '联系方式：8690-2484'
        self.label_ps.setText(ps_txt)
        self.label_ps.setFont(QFont('SanSerif', 10))
        self.label_ps.setWordWrap(True)
        self.label_ps.setFixedSize(1470, 75)
        v_box_label_ps = QVBoxLayout()
        v_box_label_ps.addWidget(self.label_ps)

        h_box_input = QHBoxLayout()
        h_box_input.addLayout(v_box_tool_name)
        h_box_input.addLayout(v_box_care_date)
        h_box_input.addLayout(v_box_care_person)
        h_box_input.addLayout(v_box_care_cycle)
        h_box_input.addLayout(v_box_warning_days)

        v_box_alarm_button = QVBoxLayout()
        v_box_alarm_button.addLayout(h_box_label_alarm)

        v_box_input_button = QVBoxLayout()
        v_box_input_button.addLayout(h_box_input)
        v_box_input_button.addLayout(h_box_button_care_finish)

        h_box_title = QHBoxLayout()
        h_box_title.addWidget(self.label_title)

        h_box_table = QHBoxLayout()
        h_box_table.addWidget(self.table_view)

        h_box_table_all_data = QHBoxLayout()
        h_box_table_all_data.addWidget(self.table_view_all_data)

        v_box = QVBoxLayout()
        v_box.addLayout(h_box_title)
        v_box.addLayout(h_box_table)
        v_box.addLayout(v_box_alarm_button)
        v_box.addLayout(v_box_input_button)
        v_box.addLayout(v_box_label_ps)
        v_box.addLayout(h_box_table_all_data)

        self.time_select_1m()
        time_select = QtCore.QTimer(self)
        time_select.timeout.connect(self.time_select_1m)
        time_select.start(60000)

        self.setLayout(v_box)

    def time_select_1m(self):
        need_care_tool(self)

    def button_color_green(self):
        self.button_care_finish.setStyleSheet('background-color:rgb(148, 138, 84)')

    def button_color_white(self):
        self.button_care_finish.setStyleSheet('background-color:rgb(240, 240, 240), border:none')

    def care_tool_write(self):

        def data_write(row):
            sourcefile = '备品保养数据.xlsx'  # 获取数据源
            workbook = load_workbook(sourcefile)
            worksheet = workbook.active
            worksheet.cell(row, 1, str(write_list[0]))  # 写入数据到文件
            worksheet.cell(row, 2, str((write_list[1])))  # 写入数据到文件
            worksheet.cell(row, 3, str(write_list[2]))  # 写入数据到文件
            worksheet.cell(row, 4, int(write_list[3]))  # 写入数据到文件
            worksheet.cell(row, 5, str(write_list[4]))  # 写入数据到文件
            worksheet.cell(row, 6, int(write_list[5]))  # 写入数据到文件
            worksheet.cell(row, 7, int(write_list[6]))  # 写入数据到文件
            worksheet.cell(row, 8, int(write_list[7]))  # 写入数据到文件
            workbook.save(sourcefile)
            self.time_select_1m()

        if self.edit_tool_name.text() != '' and self.edit_care_date.text() != '' and len(self.edit_care_date.text()) == 10 and self.edit_care_person.text() != '' and self.edit_care_cycle.text() != '' and self.edit_warning_days.text():
            last_care_date = (datetime.date(int(str(self.edit_care_date.text())[0:4]), int(str(self.edit_care_date.text())[5:7]), int(str(self.edit_care_date.text())[8:10])) + datetime.timedelta(days=int(self.edit_care_cycle.text()))).strftime('%Y/%m/%d')
            last_care_date_remaining_days = (datetime.datetime.strptime(last_care_date, '%Y/%m/%d') - datetime.datetime.strptime((datetime.datetime.now()).strftime('%Y/%m/%d'), '%Y/%m/%d')).days
            is_alarm = int(last_care_date_remaining_days) - int(self.edit_warning_days.text())
            write_list = [self.edit_tool_name.text(), self.edit_care_date.text(), self.edit_care_person.text(), self.edit_care_cycle.text(), last_care_date, last_care_date_remaining_days, self.edit_warning_days.text(), is_alarm]
            sourcefile = './备品保养数据.xlsx'  # 获取数据源
            sourcefile_frame = pd.read_excel(sourcefile, usecols=[0], names=None)
            sourcefile_list = sourcefile_frame.values.tolist()
            tool_name_list = []
            for i in sourcefile_list:
                tool_name_list.append(i[0])

            is_data_exist = 0
            tool_exit_site = 0
            tool_no_exit_site = 0

            for tool_sum in range(0, len(tool_name_list)):
                if str(self.edit_tool_name.text()) == str(tool_name_list[tool_sum]):
                    tool_exit_site = tool_sum
                    is_data_exist = 1
                    break

                else:
                    is_data_exist = 0

            str_null = 'nan'
            for new_tool in range(0, len(tool_name_list)):
                if str(tool_name_list[new_tool]) == str(str_null):
                    tool_no_exit_site = new_tool
                    break

            if is_data_exist == 1:
                data_write(tool_exit_site + 2)
            elif is_data_exist == 0:
                data_write(tool_no_exit_site + 2)

        else:
            self.message_exit_is_null()

    def message_exit_is_null(self):
        QMessageBox.warning(self, '警告提醒', '1.输入内容禁止为空，请检查后重新输入\n2.保养日期格式为2020/05/01，请确认格式\n3.其他问题请联系：MS06H3苏晓飞 8690-2484')


def restart(app_restart, form_restart):
    form_restart.show()  # 使用show()方法将窗口控件显示在屏幕上
    if app.exec_() == 0:
        restart(app_restart, form_restart)
        sys.exit()
    sys.exit(app.exec_())  # 进入该程序的主循环;使用sys.exit()方法的退出可以保证程序完整的结束，在这种情况下系统的环境变量会记录程序是如何退出的；如果程序运行成功，exec_()的返回值为0，否则为非0


if __name__ == '__main__':

    app = QApplication(sys.argv)  # 每一个PyQt5程序中都需要有一个QApplication对象，QApplication类包含在QTWidgets模块中，sys.argv是一个命令行参数列表；Python脚本可以从Shell中执行，比如双击*.py文件，通过参数来选择启动脚本的方式
    form = MainWindow()
    form.show()  # 使用show()方法将窗口控件显示在屏幕上
    if app.exec_() == 0:
        restart(app, form)
        sys.exit()
    sys.exit(app.exec_())  # 进入该程序的主循环;使用sys.exit()方法的退出可以保证程序完整的结束，在这种情况下系统的环境变量会记录程序是如何退出的；如果程序运行成功，exec_()的返回值为0，否则为非0

    pass
